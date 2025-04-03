import openpyxl as xl
import bisect
import string
from copy import copy
import pandas as pd
from datetime import datetime

# excel file used here
excel_file = ""
monthly_file_name = ""
wb = None
wb_data = None
accounts = []
new_acc = []
current_month = ""
add = ""
current_year = 0
filename_year = 0
close = 0
cols = {"Jul":['R',1,1,7], "Aug":['S',3,2,8], "Sep":['T',5,3,9], "Oct":['U',7,4,10], "Nov":['V',9,5,11], "Dec":['W',11,6,12],
        "Jan":['X',13,7,1], "Feb":['Y',15,8,2], "Mar":['Z',17,9,3], "Apr":['AA',19,10,4], "May":['AB',21,11,5], "Jun":['AC',23,12,6],
        "1st Close":['AD',23,12,6], "2nd Close":['AF',24,12,6], "3rd Close":['AH',24,12,6], "Final Close":['AI',24,12,6]}

def eomday(month):
    days_per_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    current_year = datetime.now().year
    year = current_year
    if(month>6): 
        year = current_year + 1
    d = days_per_month[month-1]
    if month == 2 and (year % 4 == 0 and year % 100 != 0 or year % 400 == 0):
        d = 29
    return [month,d,year]
#-----------------------------------------------------------------------------STEP_1---------------------------------------------------------------------------------
def copy_monthly_sheet_data():
    global wb, wb_data, cols, current_month, close, current_year, filename_year, add
    print("Running FS Accounting Financial Macro")
    print(f"---------------------------------------")
    # load the workbook
    global wb, wb_data, current_month, close
    wb = xl.load_workbook(excel_file, keep_vba=False)
    wb_data = xl.load_workbook(excel_file, keep_vba=False, data_only=True)
    sheets = wb.sheetnames

    # Clear all data from the destination sheet
    destination_sheet = wb["FI-U227 Statement of Revenu (2"]
    destination_sheet.delete_rows(2, destination_sheet.max_row)

    # read monthly file
    #monthly_file_name = "C:\\Users\\indronil\\Downloads\\FI-U227 OCT Statement of Revenue and Expense Detail.xlsx"
    add = ""
    monthly_file = xl.load_workbook(monthly_file_name)
    current_month = monthly_file_name.split("\\")[-1][8:11].capitalize()
    if current_month=="Jun":
        if "1st close".lower() in monthly_file_name.split("\\")[-1].lower(): 
            close = 1
            add = "1st Close"
        elif "2nd close".lower() in monthly_file_name.split("\\")[-1].lower(): 
            close = 2
            add = "2nd Close"
        elif "3rd close".lower() in monthly_file_name.split("\\")[-1].lower(): 
            close = 3
            add = "3rd Close"
        else: 
            close = 4
            add = "Final Close"
    if close!=0: print("Current Month: ", current_month, " - ", add)
    else: print("Current Month: ", current_month)
    wb['SUMMARY - FS (000000)']['A5'].value = cols[current_month][2]
    if current_month=="Jun" and add != "":
        wb['SUMMARY - FS (000000)']['B5'].value = cols[add][1]
    else:
        wb['SUMMARY - FS (000000)']['B5'].value = cols[current_month][1]
    date = eomday(cols[current_month][3])
    wb['SUMMARY - FS (000000)']['B1'].value = str(date[2])
    filename_year = date[2]
    if date[0]>6: current_year = date[2]-1
    else: current_year = date[2]
    wb['SUMMARY - FS (000000)']['A4'].value = (str(date[0])+"/"+str(date[1])+"/"+str(current_year))
    source_sheet = monthly_file.worksheets[0]

    accounting_style = xl.styles.NamedStyle(name='accounting', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)')
    accounts_monthly = []
    # Copy data from the source sheet to the destination sheet
    maxm_row = source_sheet.max_row
    for row in source_sheet.iter_rows(min_row=2, max_row=maxm_row, values_only=True):
        account_number = row[6][0:6]
        try:
            accounts_monthly.append(int(account_number))
        except:
            account_number = "ACCT"

        row = list(row)
        row.insert(6,account_number)
        destination_sheet.append(row)

    print("New month's data added to 'FI-U227 Statement of Revenu (2' sheet!")
    
    accounts_monthly = list(set(accounts_monthly))
    if current_month== "Jun" and add!= "":
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx")
    else:
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx")    
    print(f"---------------------------------------")
    return accounts_monthly

#-----------------------------------------------------------------------------STEP_2---------------------------------------------------------------------------------
def compare_account_numbers():
    global new_acc, wb, cols, current_month, close, current_year, filename_year, add
    existing_accounts = []
    for cell in wb["SUMMARY - FS (000000)"]['B']:
        try:
            account_number = int(cell.value)
            if(len(str(account_number))):
                existing_accounts.append(account_number)
        except:
            account_number=0

    for acct in accounts:
        if acct not in existing_accounts:
            new_acc.append(acct)
    
    if len(new_acc) > 0:
        msg = "\nPart 1 : New accounts to add for new month" + str(new_acc)

    else:
        msg = "\nPart 1 : No new account to add for new month!"

    return msg

def compare_summary_and_others():
    summary_check = [] #data of the rows in summary to check
    check = 0 # whether to check the sheet
    msg = ""
    for sheet in wb.sheetnames:
        if sheet == "SUMMARY - FS (000000)":
            for i in range (10, len(wb[sheet]['B'])+1):
                cell = wb[sheet]['B'][i]
                if cell.value is not None:
                    summary_check.append(str(cell.value).replace(" ",""))
                elif cell.value is None and i>=10:
                    if wb[sheet]['B'][i+1].value is None and wb[sheet]['B'][i+2].value:
                        check=1
                        break
                    else:
                        summary_check.append("")
        elif sheet == "Mapping":
            check=0
            
        elif check==1:
            for i in range (9, len(wb[sheet]['B'])+1):
                cell = wb[sheet]['B'][i]
                if cell.value is not None:
                    if str(cell.value).replace(" ","") != summary_check[i-10]:
                        print("Mismatch in line ", i+1, "with summary and ", sheet, cell.value,"!=", summary_check[i-10], "!\n")
                        msg = "\nPart 2 : Checking stopped! Mismatch found in "+sheet
                        return msg
                elif cell.value is None and i>=10:
                    if wb[sheet]['B'][i+1].value is None and wb[sheet]['B'][i+2].value:
                        check=1
                        break

            print(sheet, "checking completed!")
    msg = "\nPart 2 : Individual Org checking completed succesfully!"
    print(f"---------------------------------------")
    return msg

#-----------------------------------------------------------------------------STEP_3---------------------------------------------------------------------------------
def add_account(account_number, account_name, account_type):
    global new_acc, wb, cols, current_month, close, current_year, filename_year, add
    accounts_index = {"Personnel Services":0,"Fringe Benefits":1,"Travel and Training":2,"Other Expenses":3,"Recovery":4}
    starting_row = [11]
    closing_row = []
    all_accounts = []
    sub = []
    row=10 
    break_both = False
    while True:
        if break_both:
            break
        try:
            cell = wb["SUMMARY - FS (000000)"]['B'][row]
            number = int(cell.value)
            if(len(str(number))):
                sub.append(number)
        except:
            all_accounts.append(sub)
            sub = []
            if wb["SUMMARY - FS (000000)"]['B'][row+1].value==None:
                closing_row.append(row+1)
                break_both = True
                break
            else:
                closing_row.append(row+1)
                starting_row.append(row+2)
        row += 1
    print("Start: ", starting_row, " Close: ", closing_row)

    def change_other_account_type_starting_closing(starting_row, index): # Change account type starting and closing after an insertion
        for i in range(index+1,len(starting_row)):
            starting_row[i]+=1
        for i in range(index,len(starting_row)):
            closing_row[i]+=1
        for i in range(0,len(starting_row)):
            wb["Mapping"]["O"+str(i+3)].value = starting_row[i]
            wb["Mapping"]["P"+str(i+3)].value = closing_row[i]
            wb["Mapping"]["S"+str(i+3)].value = int(wb["Mapping"]["S"+str(i+3)].value)+1
        return starting_row, closing_row
    
    def search_insert_position_and_insert(all_accounts, starting_row, index):
        accounting_style = xl.styles.NamedStyle(name='accounting', number_format='_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)')
        sheet_to_insert = wb["SUMMARY - FS (000000)"]
        insert_index = bisect.bisect_left(all_accounts[index], int(account_number))
        print("Insert at - " + str(insert_index + starting_row[index]))
        num_row = insert_index + starting_row[index]
        all_accounts[index].insert(insert_index,int(account_number))
        starting_row, closing_row = change_other_account_type_starting_closing(starting_row, index)
        insert = 0 # variable to check which sheets to insert
        for sheet_to_insert in wb.sheetnames: #inserting to all the pages
            if sheet_to_insert == "SUMMARY - FS (000000)":
                insert = 1
            elif sheet_to_insert == "Mapping":
                insert = 0
            if insert==1:
                wb[sheet_to_insert].insert_rows(num_row)
                row = str(num_row)
                wb[sheet_to_insert][num_row][0].value = account_name
                wb[sheet_to_insert][num_row][1].value = account_number
                wb[sheet_to_insert][num_row][2].value = 0
                wb[sheet_to_insert][num_row][3].value = "=C"+ row
                wb[sheet_to_insert][num_row][4].value = "=D"+ row
                wb[sheet_to_insert][num_row][5].value = "=SUM(R"+row+":AF"+row+")"
                wb[sheet_to_insert][num_row][7].value = "=+D"+row+"-F"+row+"-G"+row
                wb[sheet_to_insert][num_row][10].value = "=(((+F"+row+"-J"+row+")/$A$5)*12)+J"+row
                for r in range(11,32):
                    if r!=16:
                        wb[sheet_to_insert][num_row][r].value = 0
                for r in range(0,32):
                    if r!=16:
                        cell = wb[sheet_to_insert][num_row][r]
                        if num_row in starting_row:
                            copy1 = wb[sheet_to_insert][num_row+1][2]
                        else:
                            copy1 = wb[sheet_to_insert][num_row-1][2]
                        cell.style=copy1.style
                        cell.font=copy(copy1.font)
                        cell.fill=copy(copy1.fill)
                        cell.alignment=copy(copy1.alignment)
                        cell.border=copy(copy1.border)
                        cell.alignment = copy(copy1.alignment)
                for row_after in range(num_row+1, closing_row[6]+1):
                  if row_after not in closing_row:
                    if str(wb[sheet_to_insert][row_after][3].value).startswith('='): wb[sheet_to_insert][row_after][3].value = "=C"+ str(row_after)
                    if str(wb[sheet_to_insert][row_after][4].value).startswith('='): wb[sheet_to_insert][row_after][4].value = "=D"+ str(row_after)
                    wb[sheet_to_insert][row_after][5].value = "=SUM(R"+str(row_after)+":AF"+str(row_after)+")"
                    wb[sheet_to_insert][row_after][7].value = "=+D"+str(row_after)+"-F"+str(row_after)+"-G"+str(row_after)
                    if ")*24)" in str(wb[sheet_to_insert][row_after][10].value): wb[sheet_to_insert][row_after][10].value = "=(((+F"+str(row_after)+"-J"+str(row_after)+")/$B$5)*24)+J"+str(row_after)
                    else: wb[sheet_to_insert][row_after][10].value = "=(((+F"+str(row_after)+"-J"+str(row_after)+")/$A$5)*12)+J"+str(row_after)
                  else:
                      idx = closing_row.index(row_after)
                      for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI']:
                          wb[sheet_to_insert][col+str(row_after)].value = "=SUM("+col+str(starting_row[idx])+":"+col+str(closing_row[idx]-1)+")"
                                 
        return all_accounts, starting_row, closing_row

    if account_type=="Personnel Services": # calling insert function according to account type
        all_accounts, starting_row, closing_row = search_insert_position_and_insert(all_accounts, starting_row, 0)
    elif account_type=="Fringe Benefits":
        all_accounts, starting_row, closing_row = search_insert_position_and_insert(all_accounts, starting_row, 1)
    elif account_type=="Travel and Training":
        all_accounts, starting_row, closing_row = search_insert_position_and_insert(all_accounts, starting_row, 2)
    elif account_type=="Other Expenses":
        all_accounts, starting_row, closing_row = search_insert_position_and_insert(all_accounts, starting_row, 3)
    elif account_type=="Recovery":
        all_accounts, starting_row, closing_row = search_insert_position_and_insert(all_accounts, starting_row, 4)
    
    print("Start: ", starting_row, " Close: ", closing_row)
    print("New account "+str(account_number)+" - '"+str(account_name)+"' added successfully!") 
    try:
        new_acc.remove(int(str(account_number).replace(" ",'')))
    except:
        new_acc

    print("Account to be added: ", new_acc, "\n")
    if current_month== "Jun" and add!= "":
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx")
    else:
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx")    
#-----------------------------------------------------------------------------STEP_4---------------------------------------------------------------------------------

#def update_formula():
    sheets = wb.sheetnames
    summary = wb["SUMMARY - FS (000000)"]
    for row in range(starting_row[0], closing_row[4]+1):
        for col in range(2,32): #Number of columns=32
            val = ""
            if row not in closing_row:
                    val += "="
                    start = 0
                    for sheet in sheets:
                        if sheet == "Mapping":
                            start = 0
                        if start == 1:
                            if(col<=25):
                                col_letter = str(string.ascii_uppercase[col])
                            else:
                                col_letter = str(string.ascii_uppercase[int((col/25))-1]) + str(string.ascii_uppercase[int(col%25)-1])
                            val += "+'"+sheet+"'!"+col_letter+str(row)
                        if sheet == "SUMMARY - FS (000000)":
                            start = 1
            summary[str(col_letter)+str(row)].value=val
    if current_month== "Jun" and add!= "":
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx")
    else:
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx")    
#-----------------------------------------------------------------------------STEP_5---------------------------------------------------------------------------------
def update_monthly_expenses_into_organizations():
    global wb, wb_data, cols, current_month, close, current_year, filename_year, add
    sheets = wb.sheetnames
    account_dict = {}
    for cell in wb["SUMMARY - FS (000000)"]['B']:
        try:
            account_number = str(cell.value)
            if(len(str(account_number))):
                if account_number in account_dict.keys():
                    account_number = account_number + "_2" #----------------Just adding a 2 at the end to separate from the existing one (For recovery large budget pool)----------------
                account_dict.update({account_number: int(cell.coordinate.replace('B',''))})
        except:
            account_number=0

    #---------------- Recovery account maps ----------------
    map = ['500426','500451','500501','500581','500657','500696','500701','500711','500721','500731','500741','500746','500751','500761','500771','500776','500781','500681']
    ing = ['500695', '500450', '500745', '500730', '500770', '500750', '500780', '500700', '500580', '500775', '500760', '500740', '500500', '500720', '500651', '500755', '500200', '500654', '500656', '500425', '500201', '500201', '500669', '500765', '500002']
    #-------------------------------------------------------

    exp_dict = {}
    adj_dict = {}
    com_dict = {}

    if current_month == 'Jun':
        if close==1: close_num = "1st Close"
        elif close==2: close_num = "2nd Close"
        elif close==3: close_num = "3rd Close"
        else: close_num = "Final Close"
    else:
        close_num = ""

    for row in wb['FI-U227 Statement of Revenu (2'].iter_rows(min_row=2):
        recovery = 0
        for org in sheets:
            adj = 0
            com = 0
            
            curr_org = int((row[4].value.split('-')[0]))
            if str(curr_org) in map: #check if it is a recovery account
                curr_org = curr_org-1
                recovery = 1
            if str(curr_org) in str(org):
                ins_org = org # get the organization to insert

                if str(curr_org) in ing and row[5].value=="1800-INTERNAL SERVICE" and row[6].value == "795360" and str(curr_org)=="500760":
                    ins_amount = float(row[9].value) #----------------get the value to insert----------------
                    adj = float(row[8].value) #----------------get the adjusted budget to insert----------------
                    com = float(row[11].value) #----------------get the commitments to insert----------------
                elif str(curr_org) in ing and row[5].value!="1800-INTERNAL SERVICE": #----------------18XX not added for FS Accounting----------------
                    ins_amount = float(row[9].value) #----------------get the value to insert----------------
                    adj = float(row[8].value) #----------------get the adjusted budget to insert----------------
                    com = float(row[11].value) #----------------get the commitments to insert----------------
                elif (str(curr_org) not in ing):
                    ins_amount = float(row[9].value) #----------------get the value to insert----------------
                    adj = float(row[8].value) #----------------get the adjusted budget to insert----------------
                    com = float(row[11].value) #----------------get the commitments to insert----------------
                else:
                    ins_amount = 0
                    adj = 0
                    com = 0
                break
            else: 
                ins_org = org #----------------get the organization to insert----------------
                #print(curr_org) 

        #---------------- Current Period Expenses ---------------
        cell_num = row[6].value
        if cell_num == "799800" and recovery == 1: cell_num+= "_2"

        if current_month == 'Jun': #----------------For closing----------------
          if (ins_amount!=0):
            try:
                exp_dict[(ins_org, cols[close_num][0] + str(account_dict[cell_num]))] = exp_dict[ins_org, cols[close_num][0] + str(account_dict[cell_num])]+ ins_amount
                exp_dict[(ins_org, cols[current_month][0] + str(account_dict[cell_num]))] = exp_dict[ins_org, cols[current_month][0] + str(account_dict[cell_num])]+ ins_amount
            except:
                exp_dict.update({(ins_org, cols[close_num][0] + str(account_dict[cell_num])):ins_amount})
                exp_dict.update({(ins_org, cols[current_month][0] + str(account_dict[cell_num])):ins_amount})

            if (close_num=="2nd Close"):
                try:
                    exp_dict[(ins_org, 'AE' + str(account_dict[cell_num]))] = exp_dict[ins_org, 'AE' + str(account_dict[cell_num])]+ ins_amount
                except:
                     exp_dict.update({(ins_org, 'AE' + str(account_dict[cell_num])):ins_amount})
            elif (close_num=="3rd Close"):
                try:
                    exp_dict[(ins_org, 'AG' + str(account_dict[cell_num]))] = exp_dict[ins_org, 'AG' + str(account_dict[cell_num])]+ ins_amount
                except:
                     exp_dict.update({(ins_org, 'AG' + str(account_dict[cell_num])):ins_amount})

        else: #----------------For other months----------------
          if (ins_amount!=0):
            try:
                exp_dict[(ins_org, cols[current_month][0] + str(account_dict[cell_num]))] = exp_dict[ins_org, cols[current_month][0] + str(account_dict[cell_num])]+ ins_amount
            except:
                exp_dict.update({(ins_org, cols[current_month][0] + str(account_dict[cell_num])):ins_amount})

        #---------------- Adjusted Budgets ---------------
        if current_month != 'Jul': 
            if (adj!=0):
                try:
                    adj_dict[(ins_org,'D' + str(account_dict[cell_num]))] = adj_dict[ins_org,'D' + str(account_dict[cell_num])]+ adj
                except:
                    adj_dict.update({(ins_org,'D' + str(account_dict[cell_num])):adj})
            elif (adj==0 and ins_amount==0):
                try:
                    adj_dict[(ins_org,'D' + str(account_dict[cell_num]))] = adj_dict[ins_org,'D' + str(account_dict[cell_num])]+ adj
                except:
                    adj_dict.update({(ins_org,'D' + str(account_dict[cell_num])):adj})
        else:
            if (adj!=0):
                try:
                    adj_dict[(ins_org,'C' + str(account_dict[cell_num]))] = adj_dict[ins_org,'C' + str(account_dict[cell_num])]+ adj
                except:
                    adj_dict.update({(ins_org,'C' + str(account_dict[cell_num])):adj})
            elif (adj==0 and ins_amount==0):
                try:
                    adj_dict[(ins_org,'C' + str(account_dict[cell_num]))] = adj_dict[ins_org,'C' + str(account_dict[cell_num])]+ adj
                except:
                    adj_dict.update({(ins_org,'C' + str(account_dict[cell_num])):adj})


        #---------------- Commitments ---------------
        if (com!=0):
          try:
            com_dict[(ins_org,'G' + str(account_dict[cell_num]))] = com_dict[ins_org,'G' + str(account_dict[cell_num])]+ com
          except:
            com_dict.update({(ins_org,'G' + str(account_dict[cell_num])):com})
        #------------------------------------------------
        
    for key in exp_dict.keys():
        wb[key[0]][key[1]].value = exp_dict[key]
        if 'AE' in key[1]:
            prev_close = key[1]
            prev_close = prev_close.replace('AE','AD')
            try:
                wb[key[0]][key[1]].value = exp_dict[key] - wb[key[0]][prev_close].value
            except:
                wb[key[0]][key[1]].value = exp_dict[key]
        elif 'AG' in key[1]:
            prev_close = key[1]
            prev_close = prev_close.replace('AG','AF')
            try:
                wb[key[0]][key[1]].value = exp_dict[key] - wb[key[0]][prev_close].value
            except:
                wb[key[0]][key[1]].value = exp_dict[key]

    for key in adj_dict.keys():
        wb[key[0]][key[1]].value = adj_dict[key]
    for key in com_dict.keys():
        wb[key[0]][key[1]].value = com_dict[key]

        
    print("Accounts updated with the monthly expenses\n-------------------------------------")
    msg = "\nAccounts updated with the monthly expenses"
    if current_month== "Jun" and add!= "":
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx")
        print("COMPLETED! Please check 'New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx'\n-------------------------------------")
    else:
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx")
        print("COMPLETED! Please check 'New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx'\n-------------------------------------")
    return msg
    
#-----------------------------------------------------------------------------STEP_New_FY---------------------------------------------------------------------------------
def new_fy_start():
    global wb, wb_data, current_month, current_year, filename_year, cols, add
    starting_row = [11]
    closing_row = []
    all_accounts = []
    sub = []
    row=10 
    break_both = False
    while True:
        if break_both:
            break
        try:
            cell = wb["SUMMARY - FS (000000)"]['B'][row]
            number = int(cell.value)
            if(len(str(number))):
                sub.append(number)
        except:
            all_accounts.append(sub)
            sub = []
            if wb["SUMMARY - FS (000000)"]['B'][row+1].value==None:
                closing_row.append(row+1)
                break_both = True
                break
            else:
                closing_row.append(row+1)
                starting_row.append(row+2)
        row += 1
    print("Start: ", starting_row, " Close: ", closing_row)
    
    for i in range(5,90):
        wb['Executive Summary']['BT'+str(i)].value = wb_data['Executive Summary']['BL'+str(i)].value
        wb['Executive Summary']['BS'+str(i)].value = wb_data['Executive Summary']['BK'+str(i)].value
        wb['Executive Summary']['BR'+str(i)].value = wb_data['Executive Summary']['BJ'+str(i)].value
        wb['Executive Summary']['BQ'+str(i)].value = wb_data['Executive Summary']['BI'+str(i)].value
        wb['Executive Summary']['BP'+str(i)].value = wb_data['Executive Summary']['BH'+str(i)].value
        wb['Executive Summary']['BO'+str(i)].value = wb_data['Executive Summary']['BG'+str(i)].value
        wb['Executive Summary']['BN'+str(i)].value = wb_data['Executive Summary']['BF'+str(i)].value
        #----
        wb['Executive Summary']['BL'+str(i)].value = wb_data['Executive Summary']['BD'+str(i)].value
        wb['Executive Summary']['BK'+str(i)].value = wb_data['Executive Summary']['BC'+str(i)].value
        wb['Executive Summary']['BJ'+str(i)].value = wb_data['Executive Summary']['BB'+str(i)].value
        wb['Executive Summary']['BI'+str(i)].value = wb_data['Executive Summary']['BA'+str(i)].value
        wb['Executive Summary']['BH'+str(i)].value = wb_data['Executive Summary']['AZ'+str(i)].value
        wb['Executive Summary']['BG'+str(i)].value = wb_data['Executive Summary']['AY'+str(i)].value
        wb['Executive Summary']['BF'+str(i)].value = wb_data['Executive Summary']['AX'+str(i)].value
        #----
        wb['Executive Summary']['BD'+str(i)].value = wb_data['Executive Summary']['AV'+str(i)].value
        wb['Executive Summary']['BC'+str(i)].value = wb_data['Executive Summary']['AU'+str(i)].value
        wb['Executive Summary']['BB'+str(i)].value = wb_data['Executive Summary']['AT'+str(i)].value
        wb['Executive Summary']['BA'+str(i)].value = wb_data['Executive Summary']['AS'+str(i)].value
        wb['Executive Summary']['AZ'+str(i)].value = wb_data['Executive Summary']['AR'+str(i)].value
        wb['Executive Summary']['AY'+str(i)].value = wb_data['Executive Summary']['AQ'+str(i)].value
        wb['Executive Summary']['AX'+str(i)].value = wb_data['Executive Summary']['AP'+str(i)].value
        #----
        wb['Executive Summary']['AV'+str(i)].value = wb_data['Executive Summary']['AN'+str(i)].value
        wb['Executive Summary']['AU'+str(i)].value = wb_data['Executive Summary']['AM'+str(i)].value
        wb['Executive Summary']['AT'+str(i)].value = wb_data['Executive Summary']['AL'+str(i)].value
        wb['Executive Summary']['AS'+str(i)].value = wb_data['Executive Summary']['AK'+str(i)].value
        wb['Executive Summary']['AR'+str(i)].value = wb_data['Executive Summary']['AJ'+str(i)].value
        wb['Executive Summary']['AQ'+str(i)].value = wb_data['Executive Summary']['AI'+str(i)].value
        wb['Executive Summary']['AP'+str(i)].value = wb_data['Executive Summary']['AH'+str(i)].value
        #----
        wb['Executive Summary']['AN'+str(i)].value = wb_data['Executive Summary']['U'+str(i)].value
        wb['Executive Summary']['AM'+str(i)].value = wb_data['Executive Summary']['T'+str(i)].value
        wb['Executive Summary']['AL'+str(i)].value = wb_data['Executive Summary']['S'+str(i)].value
        wb['Executive Summary']['AK'+str(i)].value = wb_data['Executive Summary']['R'+str(i)].value
        wb['Executive Summary']['AJ'+str(i)].value = wb_data['Executive Summary']['Q'+str(i)].value
        wb['Executive Summary']['AI'+str(i)].value = wb_data['Executive Summary']['P'+str(i)].value
        wb['Executive Summary']['AH'+str(i)].value = wb_data['Executive Summary']['O'+str(i)].value

    insert = 0
    change = ['G','I','J','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    for sheet_to_insert in wb.sheetnames: #inserting to all the pages
            if sheet_to_insert == "Mapping":
                insert = 0
            if insert==1:
                df = pd.read_excel(excel_file, sheet_name=sheet_to_insert)
                for num_row in range(starting_row[0],closing_row[-1]+4):
                  if num_row not in closing_row and num_row!=(closing_row[-1]+1):
                    wb[sheet_to_insert]['P'+str(num_row)].value =  wb[sheet_to_insert]['O'+str(num_row)].value
                    wb[sheet_to_insert]['O'+str(num_row)].value =  wb[sheet_to_insert]['N'+str(num_row)].value
                    wb[sheet_to_insert]['N'+str(num_row)].value =  wb[sheet_to_insert]['M'+str(num_row)].value
                    wb[sheet_to_insert]['M'+str(num_row)].value =  wb[sheet_to_insert]['L'+str(num_row)].value
                    wb[sheet_to_insert]['L'+str(num_row)].value = df.iloc[num_row][5]
                    wb[sheet_to_insert]['C'+str(num_row)].value = 0
                    wb[sheet_to_insert]['D'+str(num_row)].value = "=C"+ str(num_row)
                    wb[sheet_to_insert]['E'+str(num_row)].value = "=D"+ str(num_row)
                    for c in change:
                        wb[sheet_to_insert][c+str(num_row)].value =  0
                
            if sheet_to_insert == "SUMMARY - FS (000000)":
                insert = 1
    if current_month== "Jun" and add!= "":
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+ " - " + add +".xlsx")
    else:
        wb.save("New_Projected FY"+str(filename_year%1000)+" Budget "+ str(cols[current_month][2]) +" "+current_month+".xlsx")    
    return ""